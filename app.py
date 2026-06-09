import io
import os
import sqlite3
from datetime import datetime, timedelta
from dotenv import load_dotenv

import streamlit as st
from openpyxl import Workbook, load_workbook
from init_sqlite_db import init_database, seed_sample_data

# Load environment variables
load_dotenv()

# Initialize database
init_database()
seed_sample_data()

BANKS = ["HSBC UAE", "EIB", "ENBD", "DIB"]
BANK_HEADERS = [
    "ACC NUMBER",
    "DEBTOR NAME",
    "STATUS CODE",
    "CLAIM PAID AMOUNT",
    "CLAIM PAID DATE",
    "DATE IMPORT",
]
STATUS_CODE = "PAYMENT FILE"

MASTERLIST_TYPES = ["NEW ENDO", "PULL OUT", "MASTERLIST"]
PLACEMENTS = ["SPMAN", "SPMAW", "SPQA"]
DATE_HEADER_CANDIDATES = [
    "date",
    "claim_paid_date",
    "claim paid date",
    "date import",
    "date imported",
    "report date",
    "posting date",
]
PLACEMENT_HEADER_CANDIDATES = [
    "placement",
    "plc",
    "plc_name",
    "branch",
]
TYPE_HEADER_CANDIDATES = [
    "type",
    "masterlist type",
    "transaction type",
    "category",
]
SUM_HEADER_CANDIDATES = [
    "amount",
    "claim_paid_amount",
    "claim paid amount",
    "total",
    "sum",
    "paid_amount",
]

DIB_MASTERLIST_HEADERS = [
    "CIF",
    "NAME",
    "NATIONALITY",
    "SHADOW_ACNO",
    "CS_PRODUCT",
    "PROD_TYPE",
    "LGL_CRM",
    "GROSS_OS",
    "DPD",
    "EMAIL",
    "V_PASSPORT_NO",
    "D_DOB",
    "V_PERMANENT_NATIONAL_ID",
    "V_EMAIL_ID",
    "V_EMIRATE",
    "V_MOBILE_NO",
    "V_EMPLOYER_NAME",
    "HOME_COUNTRY_NUMBER",
    "REFERENCE_NAME",
    "REF_MOBILE",
    "REF_PHONE1",
    "REF_PHONE2",
    "CP_AMT_PAY",
    "CP_PAY_DATE",
    "N_YEAR_MODEL",
    "V_MODEL_NAME",
    "V_COLOR",
    "V_REGISTRATION_NAME",
    "V_REGISTRATION_NO",
    "V_ENGINE_NO",
    "V_CHASSIS_NO",
    "V_VENDOR_NAME",
]

DIB_SHEET_NAME_CANDIDATES = {
    "alloc": ["alloc", "allocation"],
    "customer_info": ["customer_info", "customer info", "customerinfo", "costumer_info", "costumer info"],
    "last_payment": ["last_payment", "last payment", "lastpayment"],
    "corp_details": ["corp_details", "corp details", "corpdetails"],
    "auto": ["auto"],
}

# Database Configuration
DATABASE_FILE = "masterlist.db"
UPLOAD_HISTORY_DIR = os.path.join(os.path.dirname(__file__), "uploads", "masterlist_history")

def get_db_connection():
    """Create a new database connection to SQLite for the current thread."""
    try:
        # Create a new connection for this thread (don't cache)
        # check_same_thread=False allows the connection to be used across threads,
        # but we avoid this by creating fresh connections per query
        conn = sqlite3.connect(DATABASE_FILE, timeout=10)
        conn.row_factory = sqlite3.Row
        return conn
    except Exception as e:
        st.error(f"Database connection failed: {e}")
        return None


def query_users_today():
    """Fetch user login data for today."""
    conn = get_db_connection()
    if not conn:
        return None
    
    try:
        cursor = conn.cursor()
        query = """
        SELECT 
            username AS 'USER NAME',
            name AS 'AGENT NAME',
            strftime('%H:%M:%S', first_login) AS 'LOG IN TIME',
            CASE
                WHEN strftime('%H:%M:%S', first_login) > '13:01:00' THEN 'LATE'
                ELSE 'NOT LATE'
            END AS 'LATE/NOT'
        FROM user
        WHERE DATE(first_login) = DATE('now')
        ORDER BY first_login DESC
        """
        cursor.execute(query)
        users = cursor.fetchall()
        cursor.close()
        result = [dict(row) for row in users] if users else []
        return result
    except Exception as e:
        st.error(f"Query error: {e}")
        return None
    finally:
        if conn:
            conn.close()


def query_monitoring_data(bank):
    """Fetch monitoring data for a specific bank."""
    conn = get_db_connection()
    if not conn:
        return None
    
    try:
        cursor = conn.cursor()
        query = """
        SELECT
            acc_number AS 'Account ID',
            debtor_name AS 'Debtor Name',
            status AS 'Status',
            claim_paid_amount AS 'Amount',
            date_import AS 'Date Imported',
            strftime('%H:%M:%S', last_updated) AS 'Last Updated',
            details AS 'Details'
        FROM masterlist_data
        WHERE bank = ?
        AND DATE(date_import) = DATE('now')
        LIMIT 100
        """
        cursor.execute(query, (bank,))
        data = cursor.fetchall()
        cursor.close()
        result = [dict(row) for row in data] if data else []
        return result
    except Exception as e:
        st.error(f"Query error: {e}")
        return None
    finally:
        if conn:
            conn.close()


def set_page_style():
    st.set_page_config(
        page_title="MC22 FILE",
        page_icon="💗",
        layout="wide",
    )

    st.markdown(
        """
        <style>
            /* Main Background - Dark Blue */
            .stApp, .css-18e3th9 {
                background: #0A1E3D !important;
            }

            /* Tab Navigation */
            .css-1v0mbdj button {
                color: #FFFFFF !important;
                font-weight: 600 !important;
            }
            
            /* Active Tab Styling */
            .css-1v0mbdj button[aria-selected="true"] {
                color: #2196F3 !important;
                border-bottom: 3px solid #2196F3 !important;
                border-radius: 0 0 4px 4px !important;
            }

            /* Tab Container Padding */
            .css-1v0mbdj {
                padding: 1rem 0 !important;
            }

            /* Headers */
            .stMarkdown h1, .stMarkdown h2, .stMarkdown h3 {
                color: #FFFFFF !important;
                font-weight: 700 !important;
            }

            /* Body Text */
            .stMarkdown p, .css-1d391kg .stMarkdown p, body {
                color: #FFFFFF !important;
                line-height: 1.6 !important;
                font-weight: 500 !important;
            }

            /* All text elements */
            div, span, label, p, li {
                color: #FFFFFF !important;
            }

            /* Metric Containers - White Card Style */
            .css-1o4mhx7, .metric-container {
                background-color: #FFFFFF !important;
                border: 1px solid rgba(33, 150, 243, 0.15) !important;
                border-radius: 8px !important;
                box-shadow: 0 2px 8px rgba(0, 0, 0, 0.2) !important;
                padding: 1.5rem !important;
            }

            /* Metric Label */
            .css-1o4mhx7 label, .metric-label {
                color: #666666 !important;
                font-size: 0.875rem !important;
                font-weight: 600 !important;
                margin-bottom: 0.75rem !important;
                display: block !important;
            }

            /* Metric Value - Large Bold Numbers */
            .css-1o4mhx7 div, .metric-value {
                color: #0A1E3D !important;
                font-size: 2.5rem !important;
                font-weight: 700 !important;
            }

            /* Column Containers */
            .css-1d391kg {
                background-color: transparent !important;
                border: none !important;
                box-shadow: none !important;
            }

            /* Buttons */
            .stButton>button {
                background-color: #2196F3 !important;
                color: #FFFFFF !important;
                border: none !important;
                border-radius: 8px !important;
                padding: 0.75rem 1.5rem !important;
                font-weight: 600 !important;
                transition: all 0.3s ease !important;
            }

            .stButton>button:hover {
                background-color: #1976D2 !important;
                transform: translateY(-2px) !important;
                box-shadow: 0 6px 16px rgba(33, 150, 243, 0.3) !important;
            }

            /* Input Fields */
            .stTextInput>div>div>input,
            .stSelectbox>div>div>select,
            .stFileUploader {
                background-color: #1A2F52 !important;
                border: 1px solid rgba(33, 150, 243, 0.3) !important;
                border-radius: 8px !important;
                color: #FFFFFF !important;
                padding: 0.75rem !important;
            }

            .stTextInput label,
            .stSelectbox label,
            .stFileUploader label {
                color: #FFFFFF !important;
                font-weight: 600 !important;
            }

            /* Expanders */
            .css-fk8e5p {
                background-color: #1A2F52 !important;
                border: 1px solid rgba(33, 150, 243, 0.2) !important;
                border-radius: 8px !important;
                margin: 0.75rem 0 !important;
            }

            .css-fk8e5p summary {
                color: #FFFFFF !important;
                font-weight: 600 !important;
            }

            /* Info/Success/Error Messages */
            .stAlert {
                border-radius: 8px !important;
                color: #FFFFFF !important;
            }

            .stSuccess {
                background-color: rgba(76, 175, 80, 0.15) !important;
                color: #4CAF50 !important;
                border: 1px solid rgba(76, 175, 80, 0.3) !important;
            }

            .stInfo {
                background-color: rgba(33, 150, 243, 0.15) !important;
                color: #64B5F6 !important;
                border: 1px solid rgba(33, 150, 243, 0.3) !important;
            }

            .stError {
                background-color: rgba(244, 67, 54, 0.15) !important;
                color: #EF9A9A !important;
                border: 1px solid rgba(244, 67, 54, 0.3) !important;
            }

            /* Table/Dataframe Styling */
            .stDataFrame {
                background-color: #1A2F52 !important;
            }

            .stDataFrame th {
                background-color: #0A1E3D !important;
                color: #FFFFFF !important;
                font-weight: 600 !important;
            }

            .stDataFrame td {
                background-color: #1A2F52 !important;
                color: #FFFFFF !important;
                border-color: rgba(33, 150, 243, 0.2) !important;
            }
        </style>
        """,
        unsafe_allow_html=True,
    )


def read_uploaded_workbook(uploaded_file):
    data = uploaded_file.read()
    workbook = load_workbook(io.BytesIO(data), data_only=True)
    worksheet = workbook.active
    return [[cell for cell in row] for row in worksheet.iter_rows(values_only=True)]


def format_claim_paid_date(value):
    if value is None or str(value).strip() == "":
        return ""
    if isinstance(value, datetime):
        return value.date().isoformat()

    text = str(value).strip()
    for fmt in ("%d-%b-%y", "%d-%B-%y", "%d-%b-%Y", "%d/%m/%Y", "%Y-%m-%d"):
        try:
            return datetime.strptime(text, fmt).date().isoformat()
        except Exception:
            continue
    return text


def build_formatted_workbook(bank, rows, import_date):
    workbook = Workbook()
    worksheet = workbook.active
    worksheet.title = "Formatted Upload"

    for col_idx, header in enumerate(BANK_HEADERS, start=1):
        cell = worksheet.cell(row=1, column=col_idx, value=header)
        cell.font = cell.font.copy(bold=True)
        worksheet.column_dimensions[cell.column_letter].width = max(18, len(header) + 2)

    if bank == "HSBC UAE":
        for row in rows[1:]:
            acc = row[1] if len(row) > 1 else None
            debtor = row[3] if len(row) > 3 else None
            amount = row[5] if len(row) > 5 else None
            paid_date_raw = row[4] if len(row) > 4 else None

            if not any([acc, debtor, amount, paid_date_raw]):
                continue

            worksheet.append([
                str(acc).strip() if acc is not None else "",
                str(debtor).strip() if debtor is not None else "",
                STATUS_CODE,
                amount if amount is not None else "",
                format_claim_paid_date(paid_date_raw),
                import_date,
            ])
    else:
        pass

    return workbook


def normalize_key(value):
    if value is None:
        return ""
    return str(value).strip().lower().replace(" ", "_").replace("-", "_").replace("/", "_")


def parse_date_cell(value):
    if value is None or str(value).strip() == "":
        return None
    if isinstance(value, datetime):
        return value.date()
    text = str(value).strip()
    for fmt in ("%d-%b-%y", "%d-%B-%y", "%d-%b-%Y", "%d/%m/%Y", "%Y-%m-%d", "%d-%m-%Y", "%d.%m.%Y"):
        try:
            return datetime.strptime(text, fmt).date()
        except Exception:
            continue
    return None


def safe_float(value):
    if value is None or str(value).strip() == "":
        return 0.0
    try:
        return float(value)
    except Exception:
        return 0.0


def find_column_index(header_row, candidates, fallback=None):
    if not header_row:
        return fallback
    normalized_candidates = {normalize_key(c) for c in candidates}
    for idx, cell in enumerate(header_row):
        if cell is None:
            continue
        if normalize_key(cell) in normalized_candidates:
            return idx
    return fallback


def build_masterlist_inventory_records(rows):
    if len(rows) < 2:
        return []

    header_row = rows[0]
    date_idx = find_column_index(header_row, DATE_HEADER_CANDIDATES, fallback=0)
    placement_idx = find_column_index(header_row, PLACEMENT_HEADER_CANDIDATES, fallback=1)
    type_idx = find_column_index(header_row, TYPE_HEADER_CANDIDATES, fallback=2)
    sum_idx = find_column_index(header_row, SUM_HEADER_CANDIDATES, fallback=8)

    records = []
    for row in rows[1:]:
        if not any(cell not in (None, "") for cell in row):
            continue

        date_value = None
        if date_idx is not None and date_idx < len(row):
            date_value = parse_date_cell(row[date_idx])
        if date_value is None and len(row) > 0:
            date_value = parse_date_cell(row[0])

        placement = ""
        if placement_idx is not None and placement_idx < len(row):
            placement = str(row[placement_idx]).strip().upper() if row[placement_idx] is not None else ""
        if not placement and len(row) > 1:
            candidate = str(row[1]).strip().upper() if row[1] is not None else ""
            if candidate in PLACEMENTS:
                placement = candidate

        type_value = ""
        if type_idx is not None and type_idx < len(row):
            type_value = str(row[type_idx]).strip().upper() if row[type_idx] is not None else ""
        if not type_value and len(row) > 2:
            candidate = str(row[2]).strip().upper() if row[2] is not None else ""
            if candidate in MASTERLIST_TYPES:
                type_value = candidate

        amount = 0.0
        if sum_idx is not None and sum_idx < len(row):
            amount = safe_float(row[sum_idx])

        records.append({
            "date": date_value,
            "placement": placement,
            "type": type_value,
            "sum": amount,
        })

    return records


def build_inventory_summary(records, selected_date=None, selected_placement=None, selected_type=None):
    filtered = []
    for record in records:
        if selected_date and record["date"] != selected_date:
            continue
        if selected_placement and record["placement"] != selected_placement:
            continue
        if selected_type and record["type"] != selected_type:
            continue
        filtered.append(record)

    summary = {}
    for record in filtered:
        key = (
            record["date"],
            record["placement"],
            record["type"],
        )
        if key not in summary:
            summary[key] = {"count": 0, "sum": 0.0}
        summary[key]["count"] += 1
        summary[key]["sum"] += record["sum"]

    rows = []
    for (date_value, placement, type_value), values in summary.items():
        rows.append({
            "Date": date_value.isoformat() if date_value else "",
            "Placement": placement,
            "Type": type_value,
            "Count": values["count"],
            "Sum": values["sum"],
        })

    return sorted(rows, key=lambda x: (x["Date"], x["Placement"], x["Type"]))


def build_daily_masterlist_summary(records):
    summary = {}
    for record in records:
        day = record["date"].isoformat() if record["date"] else "UNKNOWN"
        type_value = record["type"] or "UNKNOWN"
        placement = record["placement"] or "UNKNOWN"
        key = (day, placement, type_value)
        if key not in summary:
            summary[key] = {"count": 0, "sum": 0.0}
        summary[key]["count"] += 1
        summary[key]["sum"] += record["sum"]

    rows = []
    for (day, type_value, placement), values in sorted(summary.items(), key=lambda item: (item[0][0], item[0][1], item[0][2])):
        rows.append({
            "Date": day,
            "Type": type_value,
            "Placement": placement,
            "Count": values["count"],
            "Total Amount": values["sum"],
        })
    return rows


def build_placement_type_table(records):
    grouped = {}
    for record in records:
        day = record["date"].isoformat() if record["date"] else "UNKNOWN"
        placement = record["placement"] or "UNKNOWN"
        type_value = record["type"] or "UNKNOWN"
        key = (day, placement, type_value)
        grouped.setdefault(key, {"count": 0, "sum": 0.0})
        grouped[key]["count"] += 1
        grouped[key]["sum"] += record["sum"]

    rows = []
    for (day, placement, type_value), values in sorted(grouped.items(), key=lambda item: (item[0][0], item[0][1], item[0][2])):
        rows.append({
            "Date": day,
            "Placement": placement,
            "Type": type_value,
            "Count": values["count"],
            "Amount": values["sum"],
        })
    return rows


def build_inventory_pivot(records, selected_date=None, selected_placement=None, selected_type=None):
    filtered = []
    for record in records:
        if selected_date and record["date"] != selected_date:
            continue
        if selected_placement and record["placement"] != selected_placement:
            continue
        if selected_type and record["type"] != selected_type:
            continue
        filtered.append(record)

    pivot = {}
    for record in filtered:
        row_key = (
            record["date"],
            record["type"] or "",
        )
        if row_key not in pivot:
            pivot[row_key] = {placement: {"count": 0, "sum": 0.0} for placement in PLACEMENTS}
        placement = record["placement"]
        if placement not in pivot[row_key]:
            placement = placement if placement in PLACEMENTS else None
        if placement:
            pivot[row_key][placement]["count"] += 1
            pivot[row_key][placement]["sum"] += record["sum"]

    rows = []
    for (date_value, type_value), placement_data in sorted(pivot.items(), key=lambda item: ((item[0][0] or datetime.min), item[0][1] or "")):
        row = {
            "Date": date_value.isoformat() if date_value else "",
            "Type": type_value,
        }
        for placement in PLACEMENTS:
            row[f"{placement} Count"] = placement_data[placement]["count"]
            row[f"{placement} Sum"] = placement_data[placement]["sum"]
        rows.append(row)
    return rows


def find_sheet(workbook, candidates):
    normalized_candidates = {normalize_key(name) for name in candidates}
    for worksheet in workbook.worksheets:
        sheet_name = normalize_key(worksheet.title)
        if sheet_name in normalized_candidates:
            return worksheet
    return None


def load_sheet_rows(workbook, candidates):
    sheet = find_sheet(workbook, candidates)
    if not sheet:
        return []
    return [[cell for cell in row] for row in sheet.iter_rows(values_only=True)]


def build_header_index(header_row):
    if not header_row:
        return {}
    return {normalize_key(cell): idx for idx, cell in enumerate(header_row) if cell is not None}


def row_to_dict(row, header_index):
    return {
        header: row[idx] if idx < len(row) else None
        for header, idx in header_index.items()
    }


def first_match(source_dicts, key_candidates):
    for source_dict in source_dicts:
        if not source_dict:
            continue
        for key in key_candidates:
            value = source_dict.get(normalize_key(key))
            if value not in (None, ""):
                return value
    return ""


def get_value_from_source(source, key_candidates):
    if not source:
        return ""

    raw_row = source.get("row")
    header_index = source.get("header_index", {})
    for key in key_candidates:
        normalized = normalize_key(key)
        if normalized in header_index:
            idx = header_index[normalized]
            if idx < len(raw_row):
                value = raw_row[idx]
                if value is not None:
                    return value
    return ""


def get_value_from_sources(sources, key_candidates):
    for source in sources:
        value = get_value_from_source(source, key_candidates)
        if value not in (None, ""):
            return value
    return ""


def get_value_by_positions(row, positions):
    for pos in positions:
        if pos < len(row):
            value = row[pos]
            if value is not None and str(value).strip() != "":
                return value
    return ""


def build_lookup_map(rows, key_candidates, fallback_index=0):
    if len(rows) < 1:
        return {}
    header_index = build_header_index(rows[0])
    key_index = None
    for candidate in key_candidates:
        normalized = normalize_key(candidate)
        if normalized in header_index:
            key_index = header_index[normalized]
            break
    if key_index is None:
        key_index = fallback_index

    lookup = {}
    for row in rows[1:]:
        if key_index >= len(row):
            continue
        key = normalize_key(row[key_index])
        if key:
            lookup[key] = {
                "row": row,
                "header_index": header_index,
            }
    return lookup


def ensure_upload_history_folder():
    os.makedirs(UPLOAD_HISTORY_DIR, exist_ok=True)


def save_masterlist_upload_history(uploaded_file, record_count, total_amount, summary_rows):
    ensure_upload_history_folder()

    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    safe_name = "".join(ch if ch.isalnum() or ch in ("-", "_", ".") else "_" for ch in uploaded_file.name)
    stored_name = f"masterlist_{timestamp}_{safe_name}"
    stored_path = os.path.join(UPLOAD_HISTORY_DIR, stored_name)

    with open(stored_path, "wb") as handle:
        handle.write(uploaded_file.getvalue())

    conn = get_db_connection()
    try:
        cursor = conn.cursor()
        cursor.execute(
            """
            INSERT INTO masterlist_upload_history (file_name, stored_path, record_count, total_amount, summary_json)
            VALUES (?, ?, ?, ?, ?)
            """,
            (uploaded_file.name, stored_path, record_count, total_amount, str(summary_rows)),
        )
        conn.commit()
    finally:
        conn.close()

    return stored_path


def fetch_upload_history(limit=10):
    conn = get_db_connection()
    if not conn:
        return []
    try:
        cursor = conn.cursor()
        cursor.execute(
            """
            SELECT file_name, stored_path, uploaded_at, record_count, total_amount
            FROM masterlist_upload_history
            ORDER BY uploaded_at DESC
            LIMIT ?
            """,
            (limit,),
        )
        return [dict(row) for row in cursor.fetchall()]
    finally:
        conn.close()


def build_dib_masterlist_workbook(uploaded_file):
    data = uploaded_file.read()
    workbook = load_workbook(io.BytesIO(data), data_only=True)

    alloc_rows = load_sheet_rows(workbook, DIB_SHEET_NAME_CANDIDATES["alloc"])
    customer_rows = load_sheet_rows(workbook, DIB_SHEET_NAME_CANDIDATES["customer_info"])
    last_rows = load_sheet_rows(workbook, DIB_SHEET_NAME_CANDIDATES["last_payment"])
    corp_rows = load_sheet_rows(workbook, DIB_SHEET_NAME_CANDIDATES["corp_details"])
    auto_rows = load_sheet_rows(workbook, DIB_SHEET_NAME_CANDIDATES["auto"])

    if not alloc_rows:
        raise ValueError("Unable to locate the ALLOC sheet in the uploaded workbook.")

    customer_map = build_lookup_map(customer_rows, ["cif", "cif_no", "cif number"], fallback_index=1)
    last_map = build_lookup_map(last_rows, ["shadow_acno", "shadow_ac_no", "shadow account", "shadow acno"], fallback_index=2)
    corp_map = build_lookup_map(corp_rows, ["cif", "cif_no", "cif number"], fallback_index=0)
    auto_map = build_lookup_map(auto_rows, ["cif", "cif_no", "cif number", "v_contract_ref_no", "v contract ref no", "contract_ref_no", "contract ref no"], fallback_index=0)

    output_workbook = Workbook()
    output_worksheet = output_workbook.active
    output_worksheet.title = "DIB Masterlist"

    for col_idx, header in enumerate(DIB_MASTERLIST_HEADERS, start=1):
        cell = output_worksheet.cell(row=1, column=col_idx, value=header)
        cell.font = cell.font.copy(bold=True)
        output_worksheet.column_dimensions[cell.column_letter].width = max(18, len(header) + 2)

    alloc_header_index = build_header_index(alloc_rows[0])
    for row in alloc_rows[1:]:
        alloc_dict = row_to_dict(row, alloc_header_index)
        alloc_source = {"row": row, "header_index": alloc_header_index}
        cif = first_match([alloc_dict], ["cif", "cif_no", "cif number"]) or get_value_by_positions(row, [0])
        shadow_acno = first_match([alloc_dict], ["shadow_acno", "shadow_account", "shadow account", "shadow_ac_no"]) or get_value_by_positions(row, [5])
        cif_key = normalize_key(cif)
        shadow_key = normalize_key(shadow_acno)

        customer = customer_map.get(cif_key)
        corp = corp_map.get(cif_key)
        auto = auto_map.get(cif_key) or auto_map.get(shadow_key)
        last_payment = last_map.get(shadow_key)

        email = get_value_from_sources([customer, corp, alloc_source], ["email", "email_id", "v_email_id", "e-mail"])
        output_worksheet.append([
            cif,
            get_value_from_sources([customer, alloc_source], ["name", "debtor_name", "customer_name"]),
            get_value_from_sources([customer, alloc_source], ["nationality"]),
            shadow_acno,
            get_value_from_sources([customer, alloc_source], ["cs_product", "cs product"]),
            get_value_from_sources([customer, alloc_source], ["prod_type", "product_type"]),
            get_value_from_sources([customer, alloc_source], ["lgl_crm", "lgl crm"]),
            get_value_from_sources([customer, alloc_source], ["gross_os", "gross_os_amount", "gross_outstanding"]),
            get_value_from_sources([customer, alloc_source], ["dpd"]),
            email,
            get_value_from_sources([corp, alloc_source], ["v_passport_no", "passport_no", "v_passport_number"]),
            get_value_from_sources([corp, alloc_source], ["d_dob", "dob", "date_of_birth"]),
            get_value_from_sources([corp, alloc_source], ["v_permanent_national_id", "permanent_national_id"]),
            get_value_from_sources([corp, alloc_source], ["v_email_id", "email_id"]),
            get_value_from_sources([corp, alloc_source], ["v_emirate", "emirate"]),
            get_value_from_sources([corp, alloc_source], ["v_mobile_no", "mobile_no", "mobile"]),
            get_value_from_sources([corp, alloc_source], ["v_employer_name", "employer_name"]),
            get_value_from_sources([customer, corp, alloc_source], ["home_country_number", "home_country_no", "country_number"]),
            get_value_from_sources([customer, corp, alloc_source], ["reference_name", "ref_name"]),
            get_value_from_sources([customer, corp, alloc_source], ["ref_mobile", "ref_mob", "reference_mobile"]),
            get_value_from_sources([customer, corp, alloc_source], ["ref_phone1", "ref_phone_1"]),
            get_value_from_sources([customer, corp, alloc_source], ["ref_phone2", "ref_phone_2"]),
            get_value_from_sources([last_payment, alloc_source], ["cp_amt_pay", "amount_paid", "payment_amount"]),
            get_value_from_sources([last_payment, alloc_source], ["cp_pay_date", "payment_date"]),
            get_value_from_sources([auto, alloc_source], ["n_year_model", "year_model"]),
            get_value_from_sources([auto, alloc_source], ["v_model_name", "model_name"]),
            get_value_from_sources([auto, alloc_source], ["v_color", "color"]),
            get_value_from_sources([auto, alloc_source], ["v_registration_name", "registration_name"]),
            get_value_from_sources([auto, alloc_source], ["v_registration_no", "registration_no"]),
            get_value_from_sources([auto, alloc_source], ["v_engine_no", "engine_no"]),
            get_value_from_sources([auto, alloc_source], ["v_chassis_no", "chassis_no"]),
            get_value_from_sources([auto, alloc_source], ["v_vendor_name", "vendor_name"]),
        ])

    return output_workbook


def make_output_filename(bank, processed_on=None):
    processed_on = processed_on or datetime.now()
    processed_date = processed_on.strftime("%Y%m%d")
    last_month = (processed_on.replace(day=1) - timedelta(days=1)).strftime("%B%Y")
    safe_bank = bank.replace(" ", "_")
    return f"{safe_bank}_{processed_date}_{last_month}.xlsx"


def main():
    set_page_style()
    st.title("MC22 FILE")
    st.markdown(
        "Use this tool to format bank upload files into a standardized Excel output with a light pink theme."
    )

    tab1, tab2, tab3, tab4, tab5, tab6, tab7, tab8 = st.tabs(["DASHBOARD", "MONITORING", "MASTERLIST", "HSBC UAE MASTERLIST", "COMPILED UPLOADS", "DIB MASTERLIST", "MONITORING DATA", "ADMIN SETTINGS"])

    with tab1:
        st.header("Dashboard")
        st.write("Welcome to MC22 FILE Dashboard.")
        
        col1, col2, col3 = st.columns(3)
        with col1:
            st.metric(label="Total Uploads", value="0", label_visibility="visible")
        with col2:
            st.metric(label="Processed Files", value="0", label_visibility="visible")
        with col3:
            st.metric(label="Pending", value="0", label_visibility="visible")

    with tab2:
        st.header("Monitoring")
        st.markdown("Monitor real-time activity and processing status by bank.")
        
        bank_monitor = st.selectbox("Select Bank to Monitor", BANKS, key="monitor_bank")
        
        # Display timestamp
        col_timestamp = st.columns([1, 3])
        with col_timestamp[1]:
            current_time = datetime.now().strftime("%A, %B %d, %Y %I:%M:%S %p")
            st.caption(f"Last Updated: {current_time}")
        
        # Fetch data from SQL
        monitor_data = query_monitoring_data(bank_monitor)
        
        if monitor_data:
            # Metric cards
            col1, col2, col3, col4, col5 = st.columns(5)
            with col1:
                st.metric(label="Total Records", value=str(len(monitor_data)), label_visibility="visible")
            with col2:
                processed = sum(1 for row in monitor_data if row.get("Status") == "Processed")
                st.metric(label="Processed", value=str(processed), label_visibility="visible")
            with col3:
                pending = sum(1 for row in monitor_data if row.get("Status") == "Pending")
                st.metric(label="Pending", value=str(pending), label_visibility="visible")
            with col4:
                errors = sum(1 for row in monitor_data if row.get("Status") == "Error")
                st.metric(label="Errors", value=str(errors), label_visibility="visible")
            with col5:
                success_rate = ((len(monitor_data) - errors) / len(monitor_data) * 100) if monitor_data else 0
                st.metric(label="Success Rate", value=f"{success_rate:.1f}%", label_visibility="visible")
            
            st.markdown("---")
            
            # Data table
            st.subheader(f"Processing Details - {bank_monitor}")
            st.dataframe(monitor_data, use_container_width=True, height=300)
        else:
            st.warning("Unable to fetch data from database. Please check the database connection.")

    with tab3:
        st.header("Masterlist Processing")
        st.markdown(
            """
            Upload bank files and format them into standardized Excel output.
            """
        )

        with st.expander("How it works", expanded=True):
            st.markdown(
                """
                - Upload an Excel workbook for the selected bank.
                - For `HSBC UAE`, the output extracts:
                  - `ACC NUMBER` from column B
                  - `DEBTOR NAME` from column D
                  - `STATUS CODE` as `PAYMENT FILE`
                  - `CLAIM PAID AMOUNT` from column F
                  - `CLAIM PAID DATE` from column E, converted to `yyyy-mm-dd`
                  - `DATE IMPORT` as the current date
                - For `ENBD`, `EIB`, and `DIB`, the app currently creates the header-only workbook until extraction rules are added.
                """
            )

        bank = st.selectbox("Select Bank", BANKS)
        uploaded_file = st.file_uploader("Upload bank workbook", type=["xlsx", "xlsm"])

        if uploaded_file is not None:
            try:
                rows = read_uploaded_workbook(uploaded_file)
                import_date = datetime.now().strftime("%Y-%m-%d")
                workbook = build_formatted_workbook(bank, rows, import_date)

                output_buffer = io.BytesIO()
                workbook.save(output_buffer)
                output_buffer.seek(0)

                filename = make_output_filename(bank)
                st.success("File formatted successfully.")
                st.download_button(
                    label="Download formatted Excel",
                    data=output_buffer,
                    file_name=filename,
                    mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                )

                st.info("For the inventory-style table like the screenshot, upload the file in the HSBC UAE MASTERLIST tab.")

                if bank != "HSBC UAE":
                    st.info("Note: this bank currently returns the header-only workbook until extraction rules are configured.")
            except Exception as error:
                st.error(f"Error processing upload: {error}")

    with tab4:
        st.header("HSBC UAE Masterlist Inventory")
        st.markdown(
            """
            Select the Date, Placement, and Type before uploading the HSBC UAE masterlist file.

            After upload, the page will display the inventory table grouped by the chosen filters.
            """
        )

        col_date, col_placement, col_type = st.columns(3)
        with col_date:
            selected_date = st.date_input("Date", value=datetime.now().date(), key="hsbc_inventory_date")
        with col_placement:
            selected_placement = st.selectbox("Placement", ["All Placements"] + PLACEMENTS, key="hsbc_inventory_placement")
        with col_type:
            selected_type = st.selectbox("Type", ["All Types"] + MASTERLIST_TYPES, key="hsbc_inventory_type")

        hsbc_upload = st.file_uploader("Upload HSBC UAE masterlist workbook", type=["xlsx", "xlsm"], key="hsbc_masterlist")
        inventory_records = []

        if hsbc_upload is not None:
            try:
                rows = read_uploaded_workbook(hsbc_upload)
                inventory_records = build_masterlist_inventory_records(rows)

                if not inventory_records:
                    st.warning("No masterlist inventory records detected in the uploaded file. Please verify the worksheet layout and headers.")
            except Exception as error:
                st.error(f"Error processing HSBC UAE masterlist: {error}")

        if inventory_records:
            selected_date_filter = selected_date
            selected_placement_filter = None if selected_placement == "All Placements" else selected_placement
            selected_type_filter = None if selected_type == "All Types" else selected_type

            summary = build_inventory_pivot(
                inventory_records,
                selected_date=selected_date_filter,
                selected_placement=selected_placement_filter,
                selected_type=selected_type_filter,
            )
            daily_summary = build_daily_masterlist_summary(inventory_records)

            if summary:
                total_count = sum(row[f"{placement} Count"] for row in summary for placement in PLACEMENTS)
                total_sum = sum(row[f"{placement} Sum"] for row in summary for placement in PLACEMENTS)

                col1, col2, col3 = st.columns(3)
                with col1:
                    st.metric(label="Records Count", value=str(total_count), label_visibility="visible")
                with col2:
                    st.metric(label="Total Amount", value=f"{total_sum:,.2f}", label_visibility="visible")
                with col3:
                    st.metric(label="Daily Groups", value=str(len(summary)), label_visibility="visible")

                st.subheader("Daily Masterlist Summary")
                st.dataframe(daily_summary, use_container_width=True, height=340)

                st.subheader("Placement / Type Breakdown")
                placement_type_table = build_placement_type_table(inventory_records)
                st.dataframe(placement_type_table, use_container_width=True, height=420)

                st.subheader("SPMAW / SPMAN / SPQA by Day")
                st.dataframe(
                    [row for row in placement_type_table if row["Placement"] in PLACEMENTS],
                    use_container_width=True,
                    height=420,
                )

                st.caption("This table shows all dates up to today, separated by placement and transaction type: MASTERLIST, NEW ENDO, PULL OUT.")

                if hsbc_upload is not None:
                    stored_path = save_masterlist_upload_history(
                        hsbc_upload,
                        record_count=total_count,
                        total_amount=total_sum,
                        summary_rows=daily_summary,
                    )
                    st.success(f"Masterlist saved for re-checking at: {stored_path}")
            else:
                st.warning("No inventory records match the selected filters.")
        else:
            st.info("Upload an HSBC UAE workbook to generate the daily masterlist, NEW ENDO, PULL OUT, and placement summary tables.")

    with tab5:
        st.header("Compiled Uploads")
        st.markdown("Saved HSBC masterlist uploads for re-checking and review.")

        history = fetch_upload_history(limit=20)
        if history:
            st.dataframe(history, use_container_width=True, height=420)
            for item in history:
                st.caption(f"File: {item['file_name']} | Stored at: {item['stored_path']} | Records: {item['record_count']} | Amount: {item['total_amount']:.2f}")
        else:
            st.info("No compiled uploads have been saved yet. Upload a masterlist file in the HSBC UAE MASTERLIST tab to create history.")

    with tab6:
        st.header("DIB Masterlist Processor")
        st.markdown(
            """
            Upload a raw workbook containing ALLOC, CUSTOMER_INFO, LAST PAYMENT, CORP DETAILS, and AUTO sheets.
            The app merges DIB details using ALLOC as the main source and generates a downloadable Excel file.
            """
        )

        dib_upload = st.file_uploader("Upload raw DIB workbook", type=["xlsx", "xlsm"], key="dib_masterlist")
        if dib_upload is not None:
            try:
                dib_workbook = build_dib_masterlist_workbook(dib_upload)
                output_buffer = io.BytesIO()
                dib_workbook.save(output_buffer)
                output_buffer.seek(0)

                filename = f"DIB_MASTERLIST_{datetime.now().strftime('%Y%m%d')}.xlsx"
                st.success("DIB masterlist generated successfully.")
                st.download_button(
                    label="Download DIB masterlist Excel",
                    data=output_buffer,
                    file_name=filename,
                    mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                )
            except Exception as error:
                st.error(f"Error processing DIB masterlist: {error}")

    with tab7:
        st.header("Monitoring Data")
        st.write("View detailed monitoring data and analytics.")
        st.info("Data visualization features coming soon.")

    with tab8:
        st.header("Admin Settings")
        st.write("Manage account notifications and login settings.")

        subtab1, subtab2, subtab3 = st.tabs(["User Login Status", "Account Settings", "Activity Log"])

        with subtab1:
            st.subheader("Agent Login Status - Today")
            
            # Fetch user login data from SQL
            users_today = query_users_today()
            
            if users_today:
                st.dataframe(users_today, use_container_width=True, height=500)
                
                col_stats1, col_stats2, col_stats3 = st.columns(3)
                with col_stats1:
                    st.metric(label="Total Users", value=str(len(users_today)), label_visibility="visible")
                with col_stats2:
                    on_time_users = sum(1 for user in users_today if str(user.get("LATE/NOT", "")) == "NOT LATE")
                    st.metric(label="On Time", value=str(on_time_users), label_visibility="visible")
                with col_stats3:
                    late_users = sum(1 for user in users_today if str(user.get("LATE/NOT", "")) == "LATE")
                    st.metric(label="Late", value=str(late_users), label_visibility="visible")
            else:
                st.warning("Unable to fetch user data from database.")
        with subtab2:
            st.subheader("Account Settings")
            with st.expander("User Credentials", expanded=False):
                username_input = st.text_input("Username")
                email_input = st.text_input("Email")
                role_select = st.selectbox("Role", ["Admin", "Agent", "Supervisor", "Manager"])
                
                if st.button("Update Account"):
                    st.success("Account updated successfully.")

            with st.expander("Notification Preferences", expanded=False):
                email_notif = st.checkbox("Enable email notifications", value=True)
                login_alert = st.checkbox("Login alerts", value=True)
                error_alert = st.checkbox("Error notifications", value=True)
                
                if st.button("Save Preferences"):
                    st.success("Notification preferences saved.")

        with subtab3:
            st.subheader("Recent Activity Log")
            
            activity_log = {
                "Timestamp": ["2026-06-05 07:45:22", "2026-06-05 07:42:35", "2026-06-05 07:40:10", "2026-06-05 07:35:50", "2026-06-05 07:30:25"],
                "User": ["cgmedalla", "decajes", "cicamado", "ncessopalao", "hpbonabon"],
                "Action": ["File Uploaded", "File Processed", "Login", "File Downloaded", "Settings Updated"],
                "Details": ["HSBC UAE file uploaded", "ENBD masterlist processed", "User logged in", "Report downloaded", "Password changed"]
            }
            
            st.dataframe(activity_log, use_container_width=True)


if __name__ == "__main__":
    main()
