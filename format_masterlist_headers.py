import csv
from argparse import ArgumentParser
from pathlib import Path

from openpyxl import Workbook, load_workbook

TARGET_HEADERS = ["ACC NUMBER", "DEBTOR NAME"]


def read_rows_from_csv(path: Path):
    with path.open(mode="r", encoding="utf-8-sig", newline="") as csv_file:
        reader = csv.reader(csv_file)
        return [list(row) for row in reader]


def read_rows_from_excel(path: Path):
    workbook = load_workbook(path, data_only=True)
    worksheet = workbook.active
    return [[cell if cell is not None else "" for cell in row] for row in worksheet.iter_rows(values_only=True)]


def normalize_row(row):
    if row is None:
        return ["", ""]

    first = str(row[0]).strip() if len(row) > 0 and row[0] is not None else ""
    second = str(row[1]).strip() if len(row) > 1 and row[1] is not None else ""
    return [first, second]


def write_rows_to_csv(path: Path, rows):
    with path.open(mode="w", encoding="utf-8", newline="") as csv_file:
        writer = csv.writer(csv_file)
        writer.writerow(TARGET_HEADERS)
        for row in rows:
            writer.writerow(normalize_row(row))


def write_rows_to_excel(path: Path, rows):
    workbook = Workbook()
    worksheet = workbook.active
    worksheet.title = "Formatted Masterlist"
    worksheet.append(TARGET_HEADERS)

    for row in rows:
        worksheet.append(normalize_row(row))

    workbook.save(path)


def determine_file_type(path: Path):
    suffix = path.suffix.lower()
    if suffix == ".csv":
        return "csv"
    if suffix in {".xlsx", ".xlsm"}:
        return "xlsx"
    raise ValueError(f"Unsupported input file type: {path.suffix}")


def main():
    parser = ArgumentParser(description="Format a masterlist file with ACC NUMBER and DEBTOR NAME headers.")
    parser.add_argument("input", type=Path, help="Path to the input CSV or XLSX file.")
    parser.add_argument("output", type=Path, nargs="?", help="Optional output file path. If omitted, uses formatted_masterlist.<ext>.")
    args = parser.parse_args()

    input_path = args.input
    if not input_path.exists():
        raise FileNotFoundError(f"Input file not found: {input_path}")

    input_type = determine_file_type(input_path)
    rows = read_rows_from_csv(input_path) if input_type == "csv" else read_rows_from_excel(input_path)

    output_path = args.output
    if output_path is None:
        default_name = "formatted_masterlist.csv" if input_type == "csv" else "formatted_masterlist.xlsx"
        output_path = Path(default_name)

    output_type = determine_file_type(output_path) if output_path.suffix else input_type
    if output_path.suffix == "":
        output_path = output_path.with_suffix(".csv" if input_type == "csv" else ".xlsx")
        output_type = determine_file_type(output_path)

    if output_type == "csv":
        write_rows_to_csv(output_path, rows)
    else:
        write_rows_to_excel(output_path, rows)

    print(f"Formatted file written to: {output_path}")


if __name__ == "__main__":
    main()
