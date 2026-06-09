from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment

workbook = Workbook()

# Define sheet names and headers
sheets = [
    "HSBC UAE",
    "EIB",
    "ENBD",
    "DIB"
]

headers = [
    "Upload First",
    "Download File",
    "Mobile Numbers",
    "New Endo",
    "Pull Out",
    "Clean Masterlist"
]

# Create and format sheets
for idx, sheet_name in enumerate(sheets):
    if idx == 0:
        sheet = workbook.active
        sheet.title = sheet_name
    else:
        sheet = workbook.create_sheet(sheet_name)

    # Add header row
    for col_idx, header in enumerate(headers, start=1):
        cell = sheet.cell(row=1, column=col_idx, value=header)
        cell.font = Font(bold=True)
        cell.fill = PatternFill(start_color="FFD966", end_color="FFD966", fill_type="solid")
        cell.alignment = Alignment(horizontal="center", vertical="center")
        sheet.column_dimensions[cell.column_letter].width = max(18, len(header) + 2)

    # Add a note section for guidance
    note_row = len(headers) + 3
    sheet.cell(row=note_row, column=1, value="Instructions:")
    sheet.cell(row=note_row + 1, column=1, value="1. Paste masterlist data into the Upload First column.")
    sheet.cell(row=note_row + 2, column=1, value="2. Use Download File, Mobile Numbers, New Endo, Pull Out, and Clean Masterlist columns for processing steps.")
    sheet.cell(row=note_row + 3, column=1, value="3. Keep each bank sheet separate for clean processing.")

    for row in range(note_row, note_row + 4):
        note_cell = sheet.cell(row=row, column=1)
        note_cell.font = Font(italic=True)

filename = "MASTERLIST_PROCESSING_TEMPLATE.xlsx"
workbook.save(filename)
print(f"Created {filename}")
